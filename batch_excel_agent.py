#!/usr/bin/env python3
"""
Batch-process Excel questions via Feedcoop agent API.
"""

import argparse
import json
import sys
import time
from pathlib import Path
from typing import Optional

import requests
from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string, get_column_letter  # noqa: F401

HOST = "open.feedcoopapi.com"  # Base host for the agent API
PATH = "/agent_api/agent/chat/completion"  # Endpoint for chat completions
CONTENT_TYPE = "application/json"  # Shared content-type header for POST bodies


class AgentAPIError(RuntimeError):
    """Raised when the agent API response is invalid."""


class AgentClient:
    """
    Thin wrapper around the Feedcoop agent completion API, responsible for
    sending user prompts and streaming back the generated answers.
    """

    def __init__(self, bot_id: str, api_key: str, timeout: int = 60) -> None:
        self.bot_id = bot_id
        self.api_key = api_key
        self.timeout = timeout
        self.session = requests.Session()  # Reuse TCP connections across rows

    def complete(self, question: str, temperature: Optional[float] = None) -> str:
        """Send one question to the agent and return the streamed answer."""
        if not question:
            raise ValueError("问题内容为空")
        body = {
            "bot_id": self.bot_id,
            "messages": [{"role": "user", "content": question}],
            "stream": True,  # Stream responses so partial chunks arrive quickly
        }
        if temperature is not None:
            body["temperature"] = temperature
        return self._post(body)

    def _post(self, body: dict) -> str:
        """Perform the streaming POST request and stitch together all chunks."""
        headers = {
            "Authorization": f"Bearer {self.api_key}",
            "Content-Type": CONTENT_TYPE,
        }
        url = f"https://{HOST}{PATH}"
        with self.session.post(
            url, json=body, headers=headers, timeout=self.timeout, stream=True
        ) as resp:
            if resp.status_code != 200:
                raise AgentAPIError(f"HTTP {resp.status_code}: {resp.text[:200]}")

            chunks = []
            for line in resp.iter_lines():
                if not line:
                    continue
                line_str = line.decode("utf-8", errors="ignore").strip()
                # Non-streaming errors surface as plain text rather than SSE data
                if "invalid_request" in line_str and not line_str.startswith("data:"):
                    raise AgentAPIError(line_str)
                if not line_str.startswith("data:"):
                    continue

                data = line_str[len("data:") :].strip()
                if data in ("[DONE]", "done"):
                    break
                try:
                    payload = json.loads(data)
                except json.JSONDecodeError:
                    continue  # Ignore malformed SSE payloads

                delta = payload["choices"][0].get("delta", {})
                content = delta.get("content")
                if content:
                    chunks.append(content)

            if not chunks:
                raise AgentAPIError("未收到有效内容")
            return "".join(chunks).strip()


def process_workbook(
    client: AgentClient,
    input_path: Path,
    output_path: Path,
    sheet_name: Optional[str],
    question_column: str,
    answer_column: str,
    start_row: int,
    skip_completed: bool,
    request_interval: float,
    max_retries: int,
    retry_wait: float,
    temperature: Optional[float],
) -> dict:
    """
    Iterate over the worksheet, send each question to the agent, write answers back,
    and return counters summarizing the batch run.
    """
    wb = load_workbook(input_path, data_only=True)
    ws = wb[sheet_name] if sheet_name else wb.active

    question_col_idx = column_index_from_string(question_column.upper())
    answer_col_idx = column_index_from_string(answer_column.upper())

    processed = skipped = failures = 0
    max_row = ws.max_row

    for row in range(start_row, max_row + 1):
        cell = ws.cell(row=row, column=question_col_idx)
        question = cell.value
        if question is None or str(question).strip() == "":
            continue  # Ignore empty question rows

        question_text = str(question).strip()
        answer_cell = ws.cell(row=row, column=answer_col_idx)

        if skip_completed and answer_cell.value not in (None, ""):
            skipped += 1
            continue  # Respect existing answers

        success = False
        last_error: Optional[Exception] = None

        for attempt in range(1, max_retries + 1):
            try:
                answer = client.complete(question_text, temperature)
                answer_cell.value = answer
                processed += 1
                success = True
                print(f"[OK] Row {row} processed.")
                time.sleep(request_interval)  # Rate-limit between calls
                break
            except Exception as exc:  # pylint: disable=broad-except
                last_error = exc
                print(
                    f"[WARN] Row {row} attempt {attempt} failed: {exc}",
                    file=sys.stderr,
                )
                if attempt < max_retries:
                    time.sleep(retry_wait)  # Back off before retrying

        if not success:
            failures += 1
            answer_cell.value = f"[ERROR] {last_error}"

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)

    return {
        "processed": processed,
        "skipped": skipped,
        "failed": failures,
        "output": str(output_path),
    }


def parse_args() -> argparse.Namespace:
    """Define and parse all CLI arguments for the batch runner."""
    parser = argparse.ArgumentParser(
        description="Batch process Excel questions with the agent API."
    )
    parser.add_argument("--bot-id", required=True, help="BOT ID")
    parser.add_argument("--api-key", required=True, help="API key")
    parser.add_argument("--input", required=True, help="输入 Excel 路径")
    parser.add_argument("--output", help="输出 Excel 路径（默认生成 *_processed.xlsx）")
    parser.add_argument("--sheet-name", help="要处理的 Sheet 名称（默认激活工作表）")
    parser.add_argument("--question-column", default="A", help="问题所在列，默认 A")
    parser.add_argument("--answer-column", default="B", help="答案输出列，默认 B")
    parser.add_argument("--start-row", type=int, default=2, help="起始行，默认 2")
    parser.add_argument(
        "--skip-completed",
        action="store_true",
        help="如果答案列已有内容则跳过该行",
    )
    parser.add_argument(
        "--request-interval",
        type=float,
        default=0.2,
        help="每次调用后的等待秒数，默认 0.2",
    )
    parser.add_argument(
        "--max-retries", type=int, default=3, help="失败重试次数，默认 3 次"
    )
    parser.add_argument(
        "--retry-wait", type=float, default=2.0, help="重试前等待秒数，默认 2 秒"
    )
    parser.add_argument(
        "--temperature",
        type=float,
        default=None,
        help="可选，传给模型的 temperature",
    )
    parser.add_argument(
        "--timeout", type=int, default=60, help="HTTP 超时时间，默认 60 秒"
    )
    return parser.parse_args()


def main() -> None:
    """Entry point: parse CLI args, run the batch, summarize results."""
    args = parse_args()

    input_path = Path(args.input).expanduser().resolve()
    if not input_path.exists():
        raise SystemExit(f"未找到输入文件 {input_path}")

    if args.output:
        output_path = Path(args.output).expanduser().resolve()
    else:
        output_path = input_path.with_name(
            f"{input_path.stem}_processed{input_path.suffix}"
        )

    client = AgentClient(args.bot_id, args.api_key, timeout=args.timeout)

    stats = process_workbook(
        client=client,
        input_path=input_path,
        output_path=output_path,
        sheet_name=args.sheet_name,
        question_column=args.question_column,
        answer_column=args.answer_column,
        start_row=args.start_row,
        skip_completed=args.skip_completed,
        request_interval=args.request_interval,
        max_retries=max(1, args.max_retries),
        retry_wait=args.retry_wait,
        temperature=args.temperature,
    )

    print(
        f"完成：{stats['processed']} 条，跳过：{stats['skipped']} 条，失败：{stats['failed']} 条，结果写入 {stats['output']}"
    )


if __name__ == "__main__":
    main()
